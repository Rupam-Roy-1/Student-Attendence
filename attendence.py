from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import os
background = "#06283D"
framebg = "#EDEDED"
framefg = "#06283D"
root = Tk()
root.title("Students Attentdents System")
root.geometry("1250x700+210+100")
root.config(bg=background)

script_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(script_dir, 'Students_data.xlsx')
file = pathlib.Path(file_path)
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "SL.NO"
    sheet['B1'] = "Name"
    sheet['C1'] = "Dept"
    sheet['D1'] = "Year"
    sheet['E1'] = "Gender"
    sheet['F1'] = "Collage Id No."
    sheet['G1'] = "Voclet_Or_Regular"
    sheet['H1'] = "Mobile No"
    sheet['I1'] = "Father's Name"

    file.save(file_path)


############ Exit Window #########################
def Exit():
    root.destroy()


####################### Show Image ################################
def Showimage():
    global filename
    global img
    filename = filedialog.askopenfilename(initialdir=os.getcwd(),
                                          title="Select image file", filetype=(("JPG File", "*.jpg"),
                                                                               ("PNG File", "*.png"),
                                                                               ("ZIP File", "*.zip"),
                                                                               ("All files", "*.txt")))
    img = (Image.open(filename))
    resize_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resize_image)
    lbl.config(image=photo2)
    lbl.file = photo2



###################### SL_No..############################

def sl_no():
    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    row = sheet.max_row

    max_row_valu = sheet.cell(row=row, column=1).value

    try:
        Sl_No.set(max_row_valu + 1)

    except:
        Sl_No.set("1")


############################# Clear ################################

def Clear():
    global img
    Name.set('')
    Mobile.set('')
    Id.set('')
    Search.set('')
    Class1.set("--Select--")
    Class2.set("--Select--")
    Class3.set("--Select--")
    radio.set('')
    Fathers_name.set('')

    savebutton.config(state='normal')
    img1 = ImageTk.PhotoImage(Image.open(os.path.join(script_dir, 'Image', 'upload photo.png')))
    lbl.config(image=img1)
    lbl.image = img1
    img = ""


######################## Save ############################
def Save():
    S1 = Sl_No.get()
    D1 = Date.get()
    N1 = Name.get()
    M1 = Mobile.get()
    try:
        G1 = Gender
    except:
        messagebox.showerror("error", "Select Gender!")
    vr = Class1.get()
    collageid = Id.get()
    fathersname = Fathers_name.get()
    D2 = Class2.get()
    Y1 = Class3.get()

    if N1 == "" or M1 == "" or vr == "--Select--" or collageid == "" or fathersname == "" or Class2 == "--Select--" or Class3 == "--Select--":
        messagebox.showerror("error", "Few data is missing!")
    else:
        file = openpyxl.load_workbook('Students_data.xlsx')
        sheet = file.active
        sheet.cell(column=1, row=sheet.max_row + 1, value=S1)
        sheet.cell(column=2, row=sheet.max_row, value=N1)
        sheet.cell(column=3, row=sheet.max_row, value=D2)
        sheet.cell(column=4, row=sheet.max_row, value=Y1)
        sheet.cell(column=5, row=sheet.max_row, value=G1)
        sheet.cell(column=6, row=sheet.max_row, value=collageid)
        sheet.cell(column=7, row=sheet.max_row, value=vr)
        sheet.cell(column=8, row=sheet.max_row, value=M1)
        sheet.cell(column=9, row=sheet.max_row, value=fathersname)
        file.save(r'Students_data.xlsx')

    try:
        img.save("Student_attendence/Student Image/" + str(S1) + ".jpg")
    except:
        messagebox.showinfo("info", "Profile Picture in no available!!!!")

    messagebox.showinfo("info", "Successfully data Entered!!!")

    Clear()

    sl_no()


############################# Search #######################################
def search():
    text = Search.get()
    Clear()
    savebutton.config(state='disable')

    file = openpyxl.load_workbook(file_path)
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == int(text):
            name = row[0]
            # print(str(name))
            sl_no_position = str(name)[14:-1]
            sl_number = str(name)[15:-1]
            # print(sl_no_position)
            # print(sl_number)
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid", "Invalid Sl No !!!!!!")

    # reg_no_position showing like A2,A3,A4,...................An
    # but reg_number just showing number after A2 like 2,3,4,..............n

    x1 = sheet.cell(row=int(sl_number), column=1).value
    x2 = sheet.cell(row=int(sl_number), column=2).value
    x3 = sheet.cell(row=int(sl_number), column=3).value
    x4 = sheet.cell(row=int(sl_number), column=4).value
    x5 = sheet.cell(row=int(sl_number), column=5).value
    x6 = sheet.cell(row=int(sl_number), column=6).value
    x7 = sheet.cell(row=int(sl_number), column=7).value
    x8 = sheet.cell(row=int(sl_number), column=8).value
    x9 = sheet.cell(row=int(sl_number), column=9).value
    # x10 = sheet.cell(row=int(sl_number), column=10).value
    # x11 = sheet.cell(row=int(sl_number), column=11).value
    # x12 = sheet.cell(row=int(sl_number), column=12).value

    # print(x1)
    # print(x2)
    # print(x3)
    # print(x4)
    # print(x5)
    # print(x6)
    # print(x7)
    # print(x8)
    # print(x9)
    # print(x10)
    # print(x11)
    # print(x12)

    Sl_No.set(x1)
    Name.set(x2)
    Class2.set(x3)
    Class3.set(x4)

    if x5 == 'Female':
        R2.select()
    else:
        R1.select()
    Id.set(x6)
    Class1.set(x7)
    Mobile.set(x8)
    Fathers_name.set(x9)

    img = (Image.open("Student_attendence/Student Image/" + str(x1) + ".jpg"))
    resize_image = img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resize_image)
    lbl.config(image=photo2)
    lbl.image = photo2


######################## Update ##############################
def Update():
    S1 = Sl_No.get()
    D1 = Date.get()
    N1 = Name.get()
    M1 = Mobile.get()
    selection()
    G1 = Gender

    vr = Class1.get()
    collageid = Id.get()
    fathersname = Fathers_name.get()
    D2 = Class2.get()
    Y1 = Class3.get()

    file = openpyxl.load_workbook(file_path)
    sheet = file.active

    for row in sheet.rows:
        if row[0].value == S1:
            name = row[0]
            print(str(name))
            sl_no_position = str(name)[14:-1]
            sl_number = str(name)[15:-1]
            print(sl_number)

    sheet.cell(column=1, row=int(sl_number), value=S1)
    sheet.cell(column=2, row=int(sl_number), value=N1)
    sheet.cell(column=3, row=int(sl_number), value=D2)
    sheet.cell(column=4, row=int(sl_number), value=Y1)
    sheet.cell(column=5, row=int(sl_number), value=G1)
    sheet.cell(column=6, row=int(sl_number), value=collageid)
    sheet.cell(column=7, row=int(sl_number), value=vr)
    sheet.cell(column=8, row=int(sl_number), value=M1)
    sheet.cell(column=9, row=int(sl_number), value=fathersname)

    file.save(file_path)

    try:
        img.save("Student_attendence/Student Image/" + str(S1) + ".jpg")

    except:
        pass
    messagebox.showinfo("Update", "Update Sucessfully!!!")

    Clear()


############## Gender #######################
def selection():
    global Gender
    value = radio.get()
    if value == 1:
        Gender = "Male"

    else:
        Gender = "Female"


# TOP FRAMES
Label(root, text="Email: rupamr294@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)
Label(root, text="STUDENTS ATTENDENTS", width=10, height=2, bg="#C36464", fg="#fff", font='Algerian 20').pack(side=TOP,fill=X)

# Search box to update
Search = StringVar()
Entry(root, textvariable=Search, width=15, bd=2, font="Arial 20").place(x=820, y=70)
imageicon3 = ImageTk.PhotoImage(Image.open(os.path.join(script_dir, 'Image', 'search.png')))
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg="#68ddfa", font="Arial 13 bold",command=search)
Srch.place(x=1060, y=66)

imageicon4 = ImageTk.PhotoImage(Image.open(os.path.join(script_dir, 'Image', 'Layer_4.png')))
Update_button = Button(root, image=imageicon4, bg="#68ddfa", command=Update)
Update_button.place(x=110, y=64)

# SL__No and Date
Label(root, text="Sl_No.:", font="Arial 13", fg=framebg, bg=background).place(x=30, y=150)
Label(root, text="Date:", font="Arial 13", fg=framebg, bg=background).place(x=500, y=150)

Sl_No = IntVar()
Date = StringVar()

sl_entry = Entry(root, textvariable=Sl_No, width=15, font="Arial 12")
sl_entry.place(x=100, y=150)

sl_no()

# Date

today = date.today()
d1 = today.strftime("%d/%m/%y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=560, y=154)
Date.set(d1)

# SL_no

# Student Details
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250,
                 relief=GROOVE)
obj.place(x=30, y=200)

Label(obj, text="Student_Name :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=40)
Label(obj, text="Mobile_No:", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=80)
Label(obj, text="Voclet/regulr :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=120)
Label(obj, text="Collage ID :", font="arial 13", bg=framebg, fg=framefg).place(x=30, y=160)
Label(obj, text="Fathers Name :", font="arial 13", bg=framebg, fg=framefg).place(x=530, y=40)
Label(obj, text="Dept :", font="arial 13", bg=framebg, fg=framefg).place(x=530, y=80)
Label(obj, text="Year :", font="arial 13", bg=framebg, fg=framefg).place(x=530, y=120)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=530, y=160)

Name = StringVar()
name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=40)

Mobile = StringVar()
Mobile_entry = Entry(obj, textvariable=Mobile, width=20, font="arial 10")
Mobile_entry.place(x=160, y=80)

Class1 = Combobox(obj, values=['Voclet', 'Regular'], font="calibi 10", width=17, state="r")
Class1.place(x=160, y=120)
Class1.set("--Select--")

Id = StringVar()
Id_entry = Entry(obj, textvariable=Id, width=20, font="arial 10")
Id_entry.place(x=160, y=160)

Fathers_name = StringVar()
fname_entry = Entry(obj, textvariable=Fathers_name, width=20, font="arial 10")
fname_entry.place(x=655, y=40)

Class2 = Combobox(obj, values=['CST', 'ME', 'EE', 'CE'], font="calibi 10", width=17, state="r")
Class2.place(x=655, y=80)
Class2.set("--Select--")

Class3 = Combobox(obj, values=['1st', '2nd', '3rd'], font="calibi 10", width=17, state="r")
Class3.place(x=655, y=120)
Class3.set("--Select--")

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=650, y=160)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=700, y=160)

# image

f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = ImageTk.PhotoImage(Image.open(os.path.join(script_dir, 'Image', 'upload photo.png')))
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# Button

Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=Showimage).place(x=1000, y=370)
savebutton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
savebutton.place(x=1000, y=450)
ResetButton = Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)
Exitbutton = Button(root, text="Exit ", width=19, height=2, font="arial 12 bold", bg="gray", command=Exit).place(x=1000, y=600)

root.mainloop()
